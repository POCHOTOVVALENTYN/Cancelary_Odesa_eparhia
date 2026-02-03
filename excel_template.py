"""
Генератор шаблона Excel для заполнения данных о священниках
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os


def create_excel_template(output_path: str = "template_priests.xlsx"):
    """
    Создание шаблона Excel файла для заполнения данных о священниках
    
    Args:
        output_path: Путь для сохранения шаблона
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Священники"
    
    # Заголовки колонок
    headers = [
        "Имя",
        "Фамилия",
        "Дата рождения",
        "Место рождения",
        "Статус",
        "Дата рукоположения",
        "Место служения",
        "Образование",
        "Последняя награда"
    ]
    
    # Стили для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Заполнение заголовков
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Примеры данных
    examples = [
        ["Иван", "Иванов", "01.01.1980", "Одесса", "Протоиерей", "15.05.2005", 
         "Свято-Успенский кафедральный собор", "Одесская духовная семинария", "Наперсный крест"],
        ["Петр", "Петров", "15.03.1975", "Киев", "Иерей", "20.06.2000", 
         "Храм Святого Николая", "Киевская духовная академия", "Камилавка"],
    ]
    
    # Заполнение примеров
    example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.fill = example_fill
    
    # Настройка ширины колонок
    column_widths = {
        'A': 15,  # Имя
        'B': 15,  # Фамилия
        'C': 18,  # Дата рождения
        'D': 20,  # Место рождения
        'E': 15,  # Статус
        'F': 20,  # Дата рукоположения
        'G': 35,  # Место служения
        'H': 30,  # Образование
        'I': 25,  # Последняя награда
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Замораживание первой строки
    ws.freeze_panes = 'A2'
    
    # Создание листа с инструкциями
    ws_instructions = wb.create_sheet("Инструкция")
    
    instructions = [
        "ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ",
        "",
        "1. Заполните данные о священниках в листе 'Священники'",
        "2. Удалите примеры данных (строки 2-3) перед заполнением",
        "",
        "ОБЯЗАТЕЛЬНЫЕ ПОЛЯ:",
        "- Имя",
        "- Фамилия",
        "- Статус (Протоиерей, Иерей, Диакон, Протодиакон)",
        "",
        "ФОРМАТЫ ДАТ:",
        "- Дата рождения: DD.MM.YYYY (например: 01.01.1980)",
        "- Дата рукоположения: DD.MM.YYYY (например: 15.05.2005)",
        "",
        "СТАТУСЫ:",
        "- Протоиерей",
        "- Иерей",
        "- Диакон",
        "- Протодиакон",
        "",
        "ПРИМЕЧАНИЯ:",
        "- Поля 'Дата рождения', 'Место рождения', 'Дата рукоположения',",
        "  'Место служения', 'Образование', 'Последняя награда' являются необязательными",
        "- Можно оставлять ячейки пустыми",
        "- После заполнения сохраните файл и отправьте боту через команду /import",
    ]
    
    for row_num, instruction in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_num, column=1)
        cell.value = instruction
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
    
    ws_instructions.column_dimensions['A'].width = 80
    
    # Сохранение файла
    wb.save(output_path)
    print(f"✓ Шаблон создан: {output_path}")
    return output_path


if __name__ == "__main__":
    template_path = create_excel_template()
    print(f"\nШаблон Excel готов к использованию!")
    print(f"Файл сохранен: {os.path.abspath(template_path)}")
