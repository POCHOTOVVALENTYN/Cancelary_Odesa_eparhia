"""
Специальный импортёр для Excel-файла формата Одесской Епархии (колонки A–K)

Структура:
 A - порядковый номер (игнорируем)
 B - Фамилия Имя Отчество (одной строкой)
 C - сан: "свящ." (иерей), "прот." (протоиерей)
 D - национальность: "укр.", "рус.", "молд." и т.п.
 E - год рождения и день тезоименитства:
     Вариант 1: 1930\\n05.09.          -> дата рождения: 05.09.1930
     Вариант 2: 1984\\n23.10.\\n08.10. -> дата рождения: 23.10.1984, именины: 08.10
 F - даты рукоположения в дьякона и священника:
     Вариант 1: 1956\\n04.08\\n05.08
         -> диакон: 04.08.1956, священник: 05.08.1956
     Вариант 2: 2008\\n29.09.\\n2013\\n28.08.
         -> диакон: 29.09.2008, священник: 28.08.2013
 G - место рождения
 H - духовное образование
 I - светское образование
 J - место служения
 K - текущая награда
"""

import logging
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple

import pandas as pd

from database import Database
from models import Priest

logger = logging.getLogger(__name__)


class LegacyExcelImporter:
    """Импортёр для специфического файла Одесской Епархии (A–K)."""

    STATUS_MAP: Dict[str, str] = {
        "свящ": "Иерей",
        "прот": "Протоиерей",
    }

    NATIONALITY_MAP: Dict[str, str] = {
        "укр": "Украинец",
        "рус": "Русский",
        "молд": "Молдованин",
    }

    def __init__(self, db: Optional[Database] = None) -> None:
        self.db = db or Database()
        self.errors: List[Dict] = []
        self.success_count: int = 0
        self.total_count: int = 0

    # ---------- Вспомогательные парсеры ----------

    @staticmethod
    def _clean_lines(cell_value) -> List[str]:
        """Разбивает значение ячейки по строкам и очищает точки/пробелы."""
        if cell_value is None or (isinstance(cell_value, float) and pd.isna(cell_value)):
            return []

        s = str(cell_value).replace("\r", "\n")
        parts = []
        for raw in s.split("\n"):
            line = raw.strip()
            if not line:
                continue
            # убираем лишние точки в конце
            line = line.strip().strip(".")
            if line:
                parts.append(line)
        return parts

    @staticmethod
    def _parse_ddmm_with_year(ddmm: str, year: str) -> Optional[date]:
        """Парсит строку вида '05.09' и год '1930' в дату 05.09.1930."""
        ddmm = ddmm.strip().strip(".")
        year = year.strip()
        try:
            return datetime.strptime(f"{ddmm}.{year}", "%d.%m.%Y").date()
        except Exception:
            return None

    @staticmethod
    def _format_ddmm(ddmm: str) -> str:
        """Нормализует день тезоименитства в формат DD.MM (строка)."""
        ddmm = ddmm.strip().strip(".")
        try:
            d = datetime.strptime(ddmm, "%d.%m")
            return d.strftime("%d.%m")
        except Exception:
            # Если не получилось распарсить, возвращаем как есть
            return ddmm

    def _parse_birth_and_name_day(self, cell_value) -> Tuple[Optional[date], str]:
        """
        Парсит колонку E: год рождения, дата рождения, день тезоименитства.

        Возвращает (birth_date, name_day_str).
        """
        lines = self._clean_lines(cell_value)
        if not lines:
            return None, ""

        # Вариант 1: год + дата рождения
        if len(lines) == 2:
            year, birth_dm = lines
            birth_date = self._parse_ddmm_with_year(birth_dm, year)
            return birth_date, ""

        # Вариант 2: год + дата рождения + день тезоименитства
        if len(lines) >= 3:
            year = lines[0]
            birth_dm = lines[1]
            name_dm = lines[2]
            birth_date = self._parse_ddmm_with_year(birth_dm, year)
            name_day = self._format_ddmm(name_dm)
            return birth_date, name_day

        # Нестандартный случай
        return None, ""

    def _parse_ordinations(self, cell_value) -> Tuple[Optional[date], Optional[date]]:
        """
        Парсит колонку F: даты рукоположения.

        Возвращает (deacon_date, priest_date).
        """
        lines = self._clean_lines(cell_value)
        if not lines:
            return None, None

        # Вариант 1: год + d/m диакон + d/m священник
        if len(lines) == 3:
            year, d1, d2 = lines
            deacon = self._parse_ddmm_with_year(d1, year)
            priest = self._parse_ddmm_with_year(d2, year)
            return deacon, priest

        # Вариант 2: год1 + d/m диакон + год2 + d/m священник
        if len(lines) >= 4:
            year1, d1, year2, d2 = lines[0], lines[1], lines[2], lines[3]
            deacon = self._parse_ddmm_with_year(d1, year1)
            priest = self._parse_ddmm_with_year(d2, year2)
            return deacon, priest

        return None, None

    def _split_full_name(self, full_name: str) -> Tuple[str, str, str]:
        """
        Разбивает 'Фамилия Имя Отчество' на (name, patronymic, surname).
        Если формат неожиданный, старается вернуть максимально разумные значения.
        """
        parts = [p for p in str(full_name).strip().split() if p]
        if len(parts) >= 3:
            surname = parts[0]
            name = parts[1]
            patronymic = " ".join(parts[2:])
        elif len(parts) == 2:
            surname, name = parts
            patronymic = ""
        elif len(parts) == 1:
            surname = parts[0]
            name = ""
            patronymic = ""
        else:
            return "", "", ""
        return name, patronymic, surname

    def _map_status(self, raw_status: str) -> str:
        s = raw_status.lower().strip()
        for key, value in self.STATUS_MAP.items():
            if key in s:
                return value
        # Если не нашли по кодам, возвращаем как есть
        return raw_status.strip()

    def _map_nationality(self, raw_nat: str) -> str:
        if not raw_nat:
            return ""
        s = raw_nat.lower().strip().strip(".")
        for key, value in self.NATIONALITY_MAP.items():
            if s.startswith(key):
                return value
        return raw_nat.strip()

    # ---------- Основная логика ----------

    def import_from_file(self, file_path: str, update_existing: bool = False) -> Dict:
        """
        Импорт из файла формата A–K.

        Колонки (индексы):
         0:A, 1:B, 2:C, 3:D, 4:E, 5:F, 6:G, 7:H, 8:I, 9:J, 10:K
        """
        self.errors = []
        self.success_count = 0
        self.total_count = 0

        # Читаем без заголовков, чтобы точно работать по индексам
        df = pd.read_excel(file_path, sheet_name=0, header=None, engine="openpyxl")
        self.total_count = len(df)
        logger.info("Legacy import: всего строк в файле: %s", self.total_count)

        for idx, row in df.iterrows():
            excel_row_num = idx + 1  # нумерация строк в Excel

            try:
                # Пропустим полностью пустые строки
                if row.isna().all():
                    continue

                # Попробуем пропустить строку-заголовок, если есть
                index_val = row.get(0)
                if isinstance(index_val, str) and not index_val.strip().isdigit():
                    # предполагаем, что это заголовок
                    continue

                full_name = row.get(1)
                if not isinstance(full_name, str) or not full_name.strip():
                    self.errors.append(
                        {
                            "row": excel_row_num,
                            "errors": ["Пустое ФИО в колонке B"],
                            "data": row.to_dict(),
                        }
                    )
                    continue

                name, patronymic, surname = self._split_full_name(full_name)
                if not surname:
                    self.errors.append(
                        {
                            "row": excel_row_num,
                            "errors": ["Не удалось распарсить ФИО"],
                            "data": row.to_dict(),
                        }
                    )
                    continue

                # Статус (сан)
                raw_status = str(row.get(2) or "").strip()
                if not raw_status:
                    self.errors.append(
                        {
                            "row": excel_row_num,
                            "errors": ["Пустой сан в колонке C"],
                            "data": row.to_dict(),
                        }
                    )
                    continue
                status = self._map_status(raw_status)

                # Национальность
                nationality = self._map_nationality(str(row.get(3) or ""))

                # Дата рождения и день тезоименитства
                birth_date, name_day = self._parse_birth_and_name_day(row.get(4))

                # Даты рукоположения
                deacon_date, priest_date = self._parse_ordinations(row.get(5))

                # Остальные поля
                birth_place = str(row.get(6) or "").strip()
                spiritual_education = str(row.get(7) or "").strip()
                secular_education = str(row.get(8) or "").strip()
                service_place = str(row.get(9) or "").strip()
                last_reward = str(row.get(10) or "").strip()

                priest = Priest(
                    name=name or "",
                    patronymic=patronymic or "",
                    surname=surname or "",
                    birth_date=birth_date,
                    birth_place=birth_place,
                    nationality=nationality,
                    status=status,
                    name_day=name_day,
                    deacon_ordination_date=deacon_date,
                    priest_ordination_date=priest_date,
                    service_place=service_place,
                    education=spiritual_education,
                    secular_education=secular_education,
                    last_reward=last_reward,
                )

                # Проверка дубликатов по ФИО (имя+фамилия+отчество)
                search_query = f"{priest.name} {priest.surname}"
                existing = self.db.search_priests(search_query)
                exact_match = any(
                    p.name.lower() == priest.name.lower()
                    and p.surname.lower() == priest.surname.lower()
                    and (p.patronymic or "").lower()
                    == (priest.patronymic or "").lower()
                    for p in existing
                )

                if exact_match and not update_existing:
                    self.errors.append(
                        {
                            "row": excel_row_num,
                            "errors": [
                                f"Священник {priest.surname} {priest.name} {priest.patronymic} уже существует"
                            ],
                            "data": row.to_dict(),
                        }
                    )
                    continue

                if exact_match and update_existing:
                    # Обновляем первую найденную запись
                    target = next(
                        p
                        for p in existing
                        if p.name.lower() == priest.name.lower()
                        and p.surname.lower() == priest.surname.lower()
                    )
                    priest.id = target.id
                    self.db.update_priest(priest)
                else:
                    self.db.add_priest(priest)

                self.success_count += 1

            except Exception as e:
                logger.exception("Ошибка при обработке строки %s: %s", excel_row_num, e)
                self.errors.append(
                    {
                        "row": excel_row_num,
                        "errors": [f"Исключение при обработке строки: {e}"],
                        "data": row.to_dict(),
                    }
                )

        result = {
            "total": self.total_count,
            "success": self.success_count,
            "errors": len(self.errors),
            "error_details": self.errors,
        }
        logger.info(
            "Legacy import завершен. Успешно: %s, Ошибок: %s",
            self.success_count,
            len(self.errors),
        )
        return result

    def get_error_report(self) -> str:
        """Формирует человекочитаемый отчёт об ошибках."""
        if not self.errors:
            return "Ошибок не обнаружено."

        report = f"Обнаружено ошибок: {len(self.errors)}\n\n"
        for error in self.errors[:20]:
            report += f"Строка {error['row']}:\n"
            for msg in error["errors"]:
                report += f"  - {msg}\n"
            report += "\n"

        if len(self.errors) > 20:
            report += f"... и ещё {len(self.errors) - 20} ошибок\n"
        return report

"""
Специальный импортёр для Excel-файла формата A–K:

A – порядковый номер
B – Фамилия Имя Отчество
C – сан (свящ., прот. и т.п.)
D – национальность (укр., рус., молд.)
E – год рождения и день тезоименитства (1 или 2 даты в ячейке)
F – даты рукоположения в диакона и священника
G – место рождения
H – духовное образование
I – светское образование
J – место служения
K – текущая награда
"""

from dataclasses import dataclass
from datetime import date
from typing import List, Dict, Optional, Tuple
import logging
import re

from openpyxl import load_workbook

from models import Priest
from database import Database
import utils

logger = logging.getLogger(__name__)


@dataclass
class ParsedDates:
    birth_date: Optional[date] = None
    name_day: str = ""  # формат DD.MM
    deacon_ordination_date: Optional[date] = None
    priest_ordination_date: Optional[date] = None


class LegacyExcelImporter:
    """Импортёр под специальный формат A–K."""

    def __init__(self, db: Optional[Database] = None):
        self.db = db or Database()
        self.errors: List[Dict] = []
        self.success_count: int = 0
        self.total_count: int = 0

    # ===== Парсинг вспомогательных полей =====

    @staticmethod
    def _split_fio(full_name: str) -> Tuple[str, str, str]:
        """
        Разбор ФИО из строки вида 'Фамилия Имя Отчество'.
        Возвращает (name, patronymic, surname).
        """
        parts = full_name.split()
        if len(parts) < 2:
            # Если структура неожиданная – считаем всё фамилией
            return full_name.strip(), "", ""

        surname = parts[0].strip()
        name = parts[1].strip()
        patronymic = " ".join(parts[2:]).strip() if len(parts) > 2 else ""
        return name, patronymic, surname

    @staticmethod
    def _map_status(raw_status: str) -> str:
        """Маппинг условных сокращений сана в статус модели."""
        if not raw_status:
            return ""
        s = raw_status.lower().replace(" ", "")
        if s.startswith("прот"):
            return "Протоиерей"
        if s.startswith("свящ"):
            return "Иерей"
        # можно расширить при необходимости
        return raw_status.strip()

    @staticmethod
    def _map_nationality(raw_nat: str) -> str:
        """Маппинг кодов национальности."""
        if not raw_nat:
            return ""
        s = raw_nat.lower().strip(". ").replace(" ", "")
        if s.startswith("укр"):
            return "Украинец"
        if s.startswith("рус"):
            return "Русский"
        if s.startswith("молд"):
            return "Молдаванин"
        return raw_nat.strip()

    @staticmethod
    def _extract_years_and_dm(text: str) -> Tuple[List[int], List[Tuple[int, int]]]:
        """
        Извлекает годы (YYYY) и пары (день, месяц) из строки с переносами.
        Возвращает (список_годов, список_(день, месяц)).
        """
        years: List[int] = []
        dms: List[Tuple[int, int]] = []

        cleaned = text.replace("\r", "\n")
        lines = [l.strip() for l in re.split(r"[\n]+", cleaned) if l.strip()]

        for line in lines:
            # Ищем год
            year_match = re.fullmatch(r"(\d{4})", line.replace(" ", ""))
            if year_match:
                years.append(int(year_match.group(1)))
                continue

            # Ищем день.месяц (возможно с точкой в конце)
            dm_match = re.fullmatch(r"(\d{1,2})\.(\d{1,2})\.?", line.replace(" ", ""))
            if dm_match:
                d = int(dm_match.group(1))
                m = int(dm_match.group(2))
                dms.append((d, m))
                continue

        return years, dms

    def _parse_birth_and_name_day(self, raw: str) -> ParsedDates:
        """
        Парсинг ячейки E: год рождения и день тезоименитства.
        Поддерживаем 2 варианта:
        1) 1930 \\n 05.09
        2) 1984 \\n 23.10 \\n 08.10
        """
        result = ParsedDates()
        if not raw:
            return result

        years, dms = self._extract_years_and_dm(str(raw))
        if years and dms:
            # Первая дата – рождение
            y = years[0]
            d1, m1 = dms[0]
            try:
                result.birth_date = date(y, m1, d1)
            except ValueError:
                pass

            # Вторая дата (если есть) – день тезоименитства (без года)
            if len(dms) > 1:
                d2, m2 = dms[1]
                result.name_day = f"{d2:02d}.{m2:02d}"

        return result

    def _parse_ordinations(self, raw: str) -> ParsedDates:
        """
        Парсинг ячейки F: даты рукоположения в диакона и священника.

        Варианты:
        1) 1956 \\n 04.08 \\n 05.08
           -> год один, два дня/месяца
        2) 2008 \\n 29.09 \\n 2013 \\n 28.08
           -> два года и две даты
        """
        result = ParsedDates()
        if not raw:
            return result

        years, dms = self._extract_years_and_dm(str(raw))

        try:
            if len(years) == 1 and len(dms) >= 1:
                y = years[0]
                d1, m1 = dms[0]
                result.deacon_ordination_date = date(y, m1, d1)
                if len(dms) > 1:
                    d2, m2 = dms[1]
                    result.priest_ordination_date = date(y, m2, d2)
            elif len(years) >= 2 and len(dms) >= 2:
                # первый год + первая дата – диакон
                y1 = years[0]
                d1, m1 = dms[0]
                result.deacon_ordination_date = date(y1, m1, d1)
                # второй год + вторая дата – священник
                y2 = years[1]
                d2, m2 = dms[1]
                result.priest_ordination_date = date(y2, m2, d2)
        except ValueError:
            # Если даты некорректные – просто оставляем None
            pass

        return result

    # ===== Основной импорт =====

    def import_from_file(self, file_path: str, update_existing: bool = False) -> Dict:
        """
        Импорт из Excel файла формата A–K.

        Args:
            file_path: путь к Excel файлу
            update_existing: обновлять существующие записи (по ФИО)
        """
        self.errors = []
        self.success_count = 0
        self.total_count = 0

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        row_index = 0

        for row in ws.iter_rows():
            row_index += 1

            # Ожидаем, что первая строка может быть заголовком — пропускаем её,
            # если в A не число.
            cell_a = row[0].value
            if row_index == 1:
                if not isinstance(cell_a, (int, float)) and not (isinstance(cell_a, str) and cell_a.strip().isdigit()):
                    continue  # считаем, что это заголовок

            # Если в A ничего нет — считаем, что данных дальше нет
            if cell_a is None:
                continue

            excel_row_number = row_index

            try:
                # B – ФИО
                full_name = str(row[1].value or "").strip()
                name, patronymic, surname = self._split_fio(full_name)

                if not name or not surname:
                    raise ValueError("Не удалось разобрать ФИО")

                # C – сан
                raw_status = str(row[2].value or "").strip()
                status = self._map_status(raw_status)
                normalized_status = utils.validate_status(status) or status

                # D – национальность
                raw_nat = str(row[3].value or "").strip()
                nationality = self._map_nationality(raw_nat)

                # E – рождение + тезоименитство
                raw_birth = row[4].value
                birth_info = self._parse_birth_and_name_day(str(raw_birth) if raw_birth is not None else "")

                # F – рукоположения
                raw_ord = row[5].value
                ord_info = self._parse_ordinations(str(raw_ord) if raw_ord is not None else "")

                # G – место рождения
                birth_place = str(row[6].value or "").strip()

                # H – духовное образование
                spiritual_education = str(row[7].value or "").strip()

                # I – светское образование
                secular_education = str(row[8].value or "").strip()

                # J – место служения
                service_place = str(row[9].value or "").strip()

                # K – текущая награда
                last_reward = str(row[10].value or "").strip() if len(row) > 10 else ""

                priest = Priest(
                    name=name,
                    patronymic=patronymic,
                    surname=surname,
                    birth_date=birth_info.birth_date,
                    birth_place=birth_place,
                    nationality=nationality,
                    status=normalized_status,
                    name_day=birth_info.name_day,
                    deacon_ordination_date=ord_info.deacon_ordination_date,
                    priest_ordination_date=ord_info.priest_ordination_date,
                    service_place=service_place,
                    education=spiritual_education,
                    secular_education=secular_education,
                    last_reward=last_reward,
                )

                # Проверка обязательных полей
                validation_errors: List[str] = []
                if not priest.name:
                    validation_errors.append("Отсутствует имя")
                if not priest.surname:
                    validation_errors.append("Отсутствует фамилия")
                if not priest.status:
                    validation_errors.append("Отсутствует статус")

                if validation_errors:
                    self.errors.append(
                        {
                            "row": excel_row_number,
                            "errors": validation_errors,
                            "data": {
                                "full_name": full_name,
                                "status": raw_status,
                                "nationality": raw_nat,
                            },
                        }
                    )
                    continue

                # Дубликаты / обновление
                existing = self.db.search_priests(f"{priest.name} {priest.surname}")
                exact_match = None
                for p in existing:
                    # сравниваем также отчество, если есть
                    if (
                        p.name.lower() == priest.name.lower()
                        and p.surname.lower() == priest.surname.lower()
                        and (p.patronymic or "").lower() == (priest.patronymic or "").lower()
                    ):
                        exact_match = p
                        break

                if exact_match and not update_existing:
                    self.errors.append(
                        {
                            "row": excel_row_number,
                            "errors": [
                                f"Священник {priest.surname} {priest.name} {priest.patronymic} уже существует"
                            ],
                            "data": {"full_name": full_name},
                        }
                    )
                    continue

                # Сохранение
                if exact_match and update_existing:
                    priest.id = exact_match.id
                    self.db.update_priest(priest)
                else:
                    self.db.add_priest(priest)

                self.success_count += 1

            except Exception as e:
                logger.error(f"Ошибка при обработке строки {excel_row_number}: {e}")
                self.errors.append(
                    {
                        "row": excel_row_number,
                        "errors": [str(e)],
                        "data": {
                            "raw_row": [cell.value for cell in row],
                        },
                    }
                )

        self.total_count = row_index

        result = {
            "total": self.total_count,
            "success": self.success_count,
            "errors": len(self.errors),
            "error_details": self.errors,
        }

        logger.info(
            f"Legacy-импорт завершен. Успешно: {self.success_count}, Ошибок: {len(self.errors)}"
        )
        return result

    def get_error_report(self) -> str:
        """Формирование текстового отчета об ошибках."""
        if not self.errors:
            return "Ошибок не обнаружено."

        report = f"Обнаружено ошибок: {len(self.errors)}\n\n"
        for error in self.errors[:20]:
            report += f"Строка {error['row']}:\n"
            for msg in error["errors"]:
                report += f"  - {msg}\n"
            report += "\n"

        if len(self.errors) > 20:
            report += f"... и ещё {len(self.errors) - 20} ошибок\n"

        return report

